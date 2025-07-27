import re
import tempfile
import pandas as pd
from openpyxl import load_workbook
import streamlit as st

import config


import uuid

def ensure_stable_id(df):
    """Ensure DataFrame has a unique 'id' column as string UUIDs."""
    if 'id' not in df.columns:
        df.insert(0, 'id', [str(uuid.uuid4()) for _ in range(len(df))])
    else:
        # Fix any missing or duplicate IDs
        existing_ids = set()
        for i in range(len(df)):
            if not df.at[i, 'id'] or df.at[i, 'id'] in existing_ids:
                df.at[i, 'id'] = str(uuid.uuid4())
            existing_ids.add(df.at[i, 'id'])
    return df

def get_strikethrough_flags_xlsx(file_path, sheet_name, header_row):
    """Reads an XLSX file and returns a matrix of strikethrough flags (monolith-style robust)."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    # Defensive: find the right sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    
    strikethrough_matrix = []

    # Defensive: only iterate as many rows as exist in the worksheet
    max_row = ws.max_row
    start_row = header_row + 2
    if start_row > max_row:
        # No data rows to process
        return []

    for row in ws.iter_rows(min_row=start_row, max_row=max_row, values_only=False):
        row_flags = []
        for cell in row:
            if cell.value is not None and hasattr(cell, "font") and cell.font and hasattr(cell.font, "strike"):
                row_flags.append("full" if cell.font.strike else "none")
            else:
                row_flags.append(None)
        strikethrough_matrix.append(row_flags)
    return strikethrough_matrix


@st.cache_data
def classify_strikethrough_row(strikes, row_cells):
    """Classifies a row as fully, partially, or not struck-through."""
    non_null_idx = [i for i, v in enumerate(row_cells) if pd.notna(v) and str(v).strip()]
    if not non_null_idx: return None
    valid_idx = [i for i in non_null_idx if i < len(strikes)]
    if not valid_idx:
        return None
    status = [strikes[i] for i in valid_idx]
    if all(s == "full" for s in status): return 1
    if any(s == "full" for s in status): return 2
    return None

def should_apply_upx_json_path(filename):
    """Checks if a filename matches the UPX Data Dictionary pattern."""
    return bool(re.match(r"UPX_Data_Dictionary_v1_\d{8}( \(\d+\))?\.xlsx", filename))

def process_upx_data_dictionary(df: pd.DataFrame) -> pd.DataFrame:
    """Applies custom forward-fill logic to build a json_path column."""
    path_columns = ["ModelName", "Primary Field", "Sub-Field (i.e. nested values)"]
    for col in path_columns:
        if col in df.columns:
            df[col] = df[col].ffill()

    def build_json_path(row):
        vals = [str(row[col]) for col in path_columns if col in row and pd.notna(row[col]) and str(row[col]).strip()]
        return ".".join(vals)

    df['json_path'] = df.apply(build_json_path, axis=1)
    cols = df.columns.tolist()
    cols.insert(0, cols.pop(cols.index('json_path')))
    return df[cols]

def embed_and_save_to_chroma(chroma_client, df, table_name, embed_cols, embedding_fn,batch_size=5000):
    df = ensure_stable_id(df)
    collection_name = f"{config.CHROMA_COLLECTION_PREFIX}_{table_name}"
    try:
        chroma_client.delete_collection(name=collection_name)
    except Exception as e:
        print(f"Collection {collection_name} does not exist or couldn't be deleted. Proceeding...")
    # Create or get existing collection
    collection = chroma_client.get_or_create_collection(
        name=collection_name,
        embedding_function=embedding_fn,
        metadata={ "hnsw:space": "cosine"}
    )

    documents, metadatas, ids = [], [], []

    for idx, row in df.iterrows():
        # Combine selected columns for embedding
        combined_text = " | ".join([str(row[col]) for col in embed_cols if pd.notna(row[col])])
        if combined_text.strip():
            documents.append(combined_text)
            stable_id = row['id']
            metadatas.append({"stable_id": stable_id, "table": table_name})
            ids.append(f"{table_name}_{stable_id}")
    total_docs = len(documents)
    # Upsert embeddings into ChromaDB (efficient insert/update)
    for i in range(0, total_docs, batch_size):
        batch_end = min(i + batch_size, total_docs)
        collection.upsert(
            documents=documents[i:batch_end],
            metadatas=metadatas[i:batch_end],
            ids=ids[i:batch_end]
        )